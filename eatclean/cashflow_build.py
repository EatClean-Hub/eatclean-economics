"""
EatClean Cash Flow Builder — FY2026

Reads three Wio bank statement CSVs (Jan/Feb/Mar 2026) from Drive,
categorizes each transaction into 2026 P&L buckets, and outputs:

  /tmp/eatclean_cashflow_2026/
    transactions.csv       — every tx categorized (ground truth)
    cashflow.json          — structured monthly + YTD summary
    Cash_Flow_2026.xlsx    — multi-sheet Excel report

Drive upload is NOT performed by this script. Run the separate
`cashflow_upload.py` after reviewing the output.

Run: python3 cashflow_build.py
"""

import sys, os, csv, json, tempfile
from collections import defaultdict
from datetime import datetime

sys.path.insert(0, os.path.dirname(__file__))

from drive_client import get_service, download_file
from cashflow_categories import categorize, BUCKET_ORDER
from wio_pdf_parser import parse_pdf

# Primary source: full Wio statement PDF covering 01/01/2026 - 24/04/2026
# Stored in /March 2026/ (a consolidated view across Jan–Apr).
# The USD account (9905287338) is inactive and not included here.
FULL_STATEMENT_PDF_ID = "1V0UjRGWC3NTAoRqvHdpVl6kfeE9iki40"
FULL_STATEMENT_LOCAL = "/tmp/eatclean_cashflow_2026/full_jan_apr_2026.pdf"

# Months to report on (Apr is partial — cutoff 2026-04-24).
MONTHS = [
    ("2026-01", "Jan 2026"),
    ("2026-02", "Feb 2026"),
    ("2026-03", "Mar 2026"),
    ("2026-04", "Apr 2026 (to 24/04)"),
]

OUT_DIR = "/tmp/eatclean_cashflow_2026"
os.makedirs(OUT_DIR, exist_ok=True)


def parse_pdf_categorized(pdf_path: str) -> list[dict]:
    """Parse the Wio statement PDF and apply categorization."""
    raw = parse_pdf(pdf_path)
    for t in raw:
        desc = f"{t['description']} {t.get('notes', '')}".strip()
        t["bucket_group"], t["bucket_line"] = categorize(desc, t["amount"])
    return raw


def compute_opening_closing(transactions: list[dict], month_key: str) -> tuple[float, float]:
    """Opening = balance-before-first-tx; Closing = balance-of-last-tx.
    Relies on `transactions` being in chronological order (PDF parse order)."""
    month_tx = [t for t in transactions if t["month"] == month_key]
    if not month_tx:
        return 0.0, 0.0
    first = month_tx[0]
    last = month_tx[-1]
    opening = (first["balance"] or 0.0) - first["amount"]
    closing = last["balance"] or 0.0
    return opening, closing


def summarize(transactions: list[dict]) -> dict:
    """Build monthly and YTD summary dict."""
    months = [m[0] for m in MONTHS]
    summary = {
        "months": months,
        "opening_balance": {},
        "closing_balance": {},
        "buckets": {},  # {group: {line: {month: total}}}
        "totals": {},   # {group: {month: total}}
    }
    for mk in months:
        op, cl = compute_opening_closing(transactions, mk)
        summary["opening_balance"][mk] = round(op, 2)
        summary["closing_balance"][mk] = round(cl, 2)

    for group, lines in BUCKET_ORDER:
        summary["buckets"][group] = {line: {m: 0.0 for m in months} for line in lines}
        summary["totals"][group] = {m: 0.0 for m in months}

    for t in transactions:
        g = t["bucket_group"]
        l = t["bucket_line"]
        m = t["month"]
        if g not in summary["buckets"]:
            summary["buckets"][g] = {}
        if l not in summary["buckets"][g]:
            summary["buckets"][g][l] = {mm: 0.0 for mm in months}
        summary["buckets"][g][l][m] = round(
            summary["buckets"][g][l][m] + t["amount"], 2
        )
        summary["totals"][g][m] = round(summary["totals"][g][m] + t["amount"], 2)

    # Net cash flow per month = sum of all bucket totals
    summary["net_cash_flow"] = {
        m: round(sum(summary["totals"][g][m] for g in summary["totals"]), 2)
        for m in months
    }
    return summary


def write_transactions_csv(transactions: list[dict], path: str):
    fields = ["month", "date", "tx_type", "description", "notes",
              "amount", "balance", "ref", "bucket_group", "bucket_line"]
    with open(path, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for t in transactions:
            w.writerow({k: t.get(k, "") for k in fields})


def write_summary_json(summary: dict, path: str):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(summary, f, indent=2, default=str)


def write_excel(summary: dict, transactions: list[dict], path: str):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()

    # --- Summary sheet ---
    ws = wb.active
    ws.title = "Cash Flow 2026"
    months = summary["months"]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    group_font = Font(bold=True)
    group_fill = PatternFill("solid", fgColor="D9E1F2")

    ws.cell(row=1, column=1, value="EatClean UAE — Cash Flow (Wio, AED)").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Generated {datetime.now().isoformat(timespec='seconds')}")

    r = 4
    headers = ["Bucket / Line"] + months + ["YTD"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    r += 1

    # Opening balance row
    ws.cell(row=r, column=1, value="Opening Balance").font = group_font
    for i, m in enumerate(months):
        ws.cell(row=r, column=2 + i, value=summary["opening_balance"][m])
    r += 2

    # Each bucket group
    for group, _lines in BUCKET_ORDER:
        if group not in summary["buckets"]:
            continue
        group_rows = summary["buckets"][group]
        # Skip group entirely if all zero
        has_any = any(
            abs(v) > 0.005
            for line in group_rows.values()
            for v in line.values()
        )
        if not has_any and group != "Revenue":
            continue

        # Group header
        c1 = ws.cell(row=r, column=1, value=group)
        c1.font = group_font
        c1.fill = group_fill
        for i, m in enumerate(months):
            v = summary["totals"].get(group, {}).get(m, 0.0)
            cell = ws.cell(row=r, column=2 + i, value=round(v, 2))
            cell.font = group_font
            cell.fill = group_fill
        ytd = sum(summary["totals"].get(group, {}).get(m, 0.0) for m in months)
        cell = ws.cell(row=r, column=2 + len(months), value=round(ytd, 2))
        cell.font = group_font
        cell.fill = group_fill
        r += 1

        for line, per_month in group_rows.items():
            # Skip lines that are all zero
            if all(abs(v) < 0.005 for v in per_month.values()):
                continue
            ws.cell(row=r, column=1, value=f"  {line}")
            for i, m in enumerate(months):
                ws.cell(row=r, column=2 + i, value=round(per_month.get(m, 0.0), 2))
            ws.cell(row=r, column=2 + len(months),
                    value=round(sum(per_month.values()), 2))
            r += 1
        r += 1

    # Net cash flow
    c1 = ws.cell(row=r, column=1, value="NET CASH FLOW")
    c1.font = Font(bold=True)
    c1.fill = PatternFill("solid", fgColor="C6EFCE")
    for i, m in enumerate(months):
        cell = ws.cell(row=r, column=2 + i, value=summary["net_cash_flow"][m])
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="C6EFCE")
    ytd_net = sum(summary["net_cash_flow"].values())
    cell = ws.cell(row=r, column=2 + len(months), value=round(ytd_net, 2))
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="C6EFCE")
    r += 1

    # Closing balance
    ws.cell(row=r, column=1, value="Closing Balance").font = group_font
    for i, m in enumerate(months):
        ws.cell(row=r, column=2 + i, value=summary["closing_balance"][m])
    r += 1

    # Column widths
    ws.column_dimensions["A"].width = 42
    for i in range(len(months) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(2 + i)].width = 14

    # --- Transactions sheet ---
    ws2 = wb.create_sheet("Transactions")
    fields = ["month", "date", "tx_type", "description", "amount",
              "balance", "bucket_group", "bucket_line", "notes", "ref"]
    for c, h in enumerate(fields, start=1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
    for i, t in enumerate(transactions, start=2):
        for c, k in enumerate(fields, start=1):
            ws2.cell(row=i, column=c, value=t.get(k))
    for col_letter, w in [("A", 9), ("B", 12), ("C", 11), ("D", 40),
                           ("E", 12), ("F", 12), ("G", 18), ("H", 28),
                           ("I", 40), ("J", 11)]:
        ws2.column_dimensions[col_letter].width = w

    # --- Uncategorized review sheet ---
    uncat = [t for t in transactions if t["bucket_group"] == "UNCATEGORIZED"]
    if uncat:
        ws3 = wb.create_sheet("Uncategorized")
        for c, h in enumerate(fields, start=1):
            cell = ws3.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
        for i, t in enumerate(uncat, start=2):
            for c, k in enumerate(fields, start=1):
                ws3.cell(row=i, column=c, value=t.get(k))
        for col_letter, w in [("A", 9), ("B", 12), ("C", 11), ("D", 40),
                               ("E", 12), ("F", 12), ("G", 18), ("H", 28),
                               ("I", 40), ("J", 11)]:
            ws3.column_dimensions[col_letter].width = w

    # --- Reconciliation sheet (Cash Flow vs 2026 P&L for Jan-Mar) ---
    try:
        from cashflow_reconcile import load_pl, CF_TO_PL
        pl = load_pl()
        ws4 = wb.create_sheet("Reconciliation")
        ws4.cell(row=1, column=1, value="Cash Flow vs 2026 P&L Reconciliation (Jan-Mar)").font = Font(bold=True, size=12)
        ws4.cell(row=2, column=1,
                 value="CF = cash paid/received. P&L = booked (accrual). Δ = CF - P&L. Negative Δ = accrued but not paid.").font = Font(italic=True)
        months_r = ["2026-01", "2026-02", "2026-03"]
        headers_r = ["CF Group", "CF Line", "P&L Label"]
        for m in months_r:
            headers_r += [f"CF {m}", f"P&L {m}", f"Δ {m}"]
        r = 4
        for c, h in enumerate(headers_r, start=1):
            cell = ws4.cell(row=r, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
        r = 5
        for (cf_group, cf_line), pl_label in CF_TO_PL.items():
            cf_per = summary["buckets"].get(cf_group, {}).get(cf_line, {})
            pl_per = pl.get(pl_label, {})
            ws4.cell(row=r, column=1, value=cf_group)
            ws4.cell(row=r, column=2, value=cf_line)
            ws4.cell(row=r, column=3, value=pl_label)
            for i, m in enumerate(months_r):
                cf_v = round(abs(cf_per.get(m, 0.0)), 2)
                pl_v = round(abs(pl_per.get(m, 0.0)), 2)
                delta = round(cf_v - pl_v, 2)
                ws4.cell(row=r, column=4 + i * 3, value=cf_v)
                ws4.cell(row=r, column=5 + i * 3, value=pl_v)
                dcell = ws4.cell(row=r, column=6 + i * 3, value=delta)
                if abs(delta) >= 1000:
                    dcell.font = Font(bold=True,
                                      color="C00000" if delta < 0 else "006100")
            r += 1
        for col_letter, w in [("A", 16), ("B", 30), ("C", 30)]:
            ws4.column_dimensions[col_letter].width = w
        for i in range(9):
            ws4.column_dimensions[openpyxl.utils.get_column_letter(4 + i)].width = 12
    except Exception as e:
        print(f"[WARN] Could not build reconciliation sheet: {e}", file=sys.stderr)

    wb.save(path)


def download_full_statement() -> str:
    if not os.path.exists(FULL_STATEMENT_LOCAL):
        print("Downloading full Wio statement (Jan-Apr 2026)...", file=sys.stderr)
        svc = get_service()
        download_file(svc, FULL_STATEMENT_PDF_ID, FULL_STATEMENT_LOCAL)
    return FULL_STATEMENT_LOCAL


def main():
    pdf_path = download_full_statement()
    all_tx = parse_pdf_categorized(pdf_path)
    # Keep only months in the MONTHS list
    month_keys = {mk for mk, _ in MONTHS}
    all_tx = [t for t in all_tx if t["month"] in month_keys]
    print(f"Parsed {len(all_tx)} transactions across {len(MONTHS)} months",
          file=sys.stderr)

    summary = summarize(all_tx)

    tx_path = os.path.join(OUT_DIR, "transactions.csv")
    json_path = os.path.join(OUT_DIR, "cashflow.json")
    xlsx_path = os.path.join(OUT_DIR, "Cash_Flow_2026.xlsx")

    write_transactions_csv(all_tx, tx_path)
    write_summary_json(summary, json_path)
    write_excel(summary, all_tx, xlsx_path)

    print("\nOutputs:")
    print(f"  {tx_path}")
    print(f"  {json_path}")
    print(f"  {xlsx_path}")

    # Quick report to stdout
    months = summary["months"]
    print("\n=== Monthly Cash Flow Summary (AED) ===")
    print(f"{'':38}" + "".join(f"{m:>14}" for m in months) + f"{'YTD':>14}")
    print(f"{'Opening Balance':38}" +
          "".join(f"{summary['opening_balance'][m]:>14,.2f}" for m in months) +
          f"{'':>14}")
    for group, _lines in BUCKET_ORDER:
        total = summary["totals"].get(group, {})
        if not any(abs(v) > 0.005 for v in total.values()):
            continue
        ytd = sum(total.get(m, 0.0) for m in months)
        print(f"{group:38}" +
              "".join(f"{total.get(m, 0.0):>14,.2f}" for m in months) +
              f"{ytd:>14,.2f}")
    print(f"{'NET CASH FLOW':38}" +
          "".join(f"{summary['net_cash_flow'][m]:>14,.2f}" for m in months) +
          f"{sum(summary['net_cash_flow'].values()):>14,.2f}")
    print(f"{'Closing Balance':38}" +
          "".join(f"{summary['closing_balance'][m]:>14,.2f}" for m in months))

    # Uncategorized tally
    uncat = [t for t in all_tx if t["bucket_group"] == "UNCATEGORIZED"]
    if uncat:
        total_uncat = sum(t["amount"] for t in uncat)
        print(f"\nUncategorized: {len(uncat)} tx, net AED {total_uncat:,.2f}")
        print("See the 'Uncategorized' sheet in the Excel file for manual review.")


if __name__ == "__main__":
    main()
