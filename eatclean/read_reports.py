"""
EatClean Accounting — Read All Reports
Connects to Drive, downloads all P&L Excel files from Report folders,
reads and outputs the data as JSON.

Run from: ~/Documents/eatclean-economics/eatclean/
Command:  python3 read_reports.py > reports_data.json
"""

import sys, os, json, tempfile
sys.path.insert(0, os.path.dirname(__file__))

from drive_client import get_service, _execute, download_file

ACCOUNTING_FOLDER_ID = "1XkmMgLanlHSwCmtAOEAn_NGI69e1VYGr"

TARGET_FILES = [
    "INCOME STATEMENT FST - PL_March 2026.xlsx",
    "INCOME STATEMENT FST - Updated_Till December 2025.xlsx",
    "Revenue analyze.xlsx",
    "CAC- FST.xlsx",
]

def find_all_report_folders(service, folder_id, results=None):
    if results is None:
        results = []
    req = service.files().list(
        q=f"'{folder_id}' in parents and mimeType='application/vnd.google-apps.folder' and trashed=false",
        fields="files(id, name)",
        pageSize=100,
    )
    resp = _execute(req)
    for f in resp.get("files", []):
        if f["name"].lower().startswith("report"):
            results.append({"id": f["id"], "name": f["name"]})
        find_all_report_folders(service, f["id"], results)
    return results

def list_files_in_folder(service, folder_id):
    req = service.files().list(
        q=f"'{folder_id}' in parents and mimeType!='application/vnd.google-apps.folder' and trashed=false",
        fields="files(id, name, mimeType)",
        pageSize=100,
    )
    resp = _execute(req)
    return resp.get("files", [])

def read_excel_to_json(path):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        result = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                if any(v is not None for v in row):
                    rows.append(list(row))
            result[sheet_name] = rows
        return result
    except Exception as e:
        return {"error": str(e)}

def main():
    print("Connecting...", file=sys.stderr)
    service = get_service()
    report_folders = find_all_report_folders(service, ACCOUNTING_FOLDER_ID)
    print(f"Found {len(report_folders)} report folders.", file=sys.stderr)

    tmp = tempfile.mkdtemp()
    output = {}

    for folder in report_folders:
        files = list_files_in_folder(service, folder["id"])
        for f in files:
            name = f["name"]
            if name not in TARGET_FILES:
                continue
            if name in output:
                continue  # already got it
            print(f"Downloading: {name}", file=sys.stderr)
            local_path = os.path.join(tmp, name)
            try:
                download_file(service, f["id"], local_path)
                if name.endswith(".xlsx"):
                    output[name] = read_excel_to_json(local_path)
                    print(f"  Done.", file=sys.stderr)
            except Exception as e:
                output[name] = {"error": str(e)}
                print(f"  Error: {e}", file=sys.stderr)

    print(json.dumps(output, indent=2, default=str))

if __name__ == "__main__":
    main()
