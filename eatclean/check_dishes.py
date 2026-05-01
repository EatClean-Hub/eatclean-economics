"""
Check all daily selection files up to 19/04/2026 for:
  - Zero-price dishes (not in price list)
  - Unrecognised dish names (not matching any known pattern)

Run from: ~/Documents/eatclean-economics/eatclean/
Command:  python3 check_dishes.py
"""

import sys, os, re, glob, tempfile
import pandas as pd
from openpyxl import load_workbook

from datetime import date, timedelta
today       = date.today()
START_DATE  = (today - timedelta(days=30)).strftime("%Y-%m-%d")
CUTOFF_DATE = today.strftime("%Y-%m-%d")

sys.path.insert(0, os.path.dirname(__file__))
from drive_client import get_service, discover_folders, download_folder_to_temp
from config import MANUAL_PRICES, VAT_DIVISOR

# ── BUILD PRICE LOOKUP ────────────────────────────────────────────────────────
print("Connecting to Drive...")
service = get_service()
folders = discover_folders(service)

print("Downloading invoice and selections...")
tmp     = tempfile.mkdtemp()
inv_tmp = os.path.join(tmp, "invoices")
sel_tmp = os.path.join(tmp, "selections")
os.makedirs(inv_tmp); os.makedirs(sel_tmp)

download_folder_to_temp(service, folders["invoices"],         inv_tmp)
download_folder_to_temp(service, folders["daily_selections"], sel_tmp)

# Build lookup from invoice Price List
lookup = {}
inv_files = glob.glob(os.path.join(inv_tmp, "**", "*.xlsx"), recursive=True)
if inv_files:
    wb = load_workbook(inv_files[0], read_only=True, data_only=True)
    if "Price List" in wb.sheetnames:
        for i, row in enumerate(wb["Price List"].iter_rows(values_only=True)):
            if i == 0: continue
            if not row[0] or not row[1]: continue
            try:
                name  = str(row[0]).strip()
                price = float(row[1]) / VAT_DIVISOR
                if name and price > 0:
                    lookup[name] = round(price, 4)
            except: continue
        print(f"  {len(lookup)} dishes loaded from price list")

# Apply MANUAL_PRICES patches
for k, v in MANUAL_PRICES.items():
    if k not in lookup:
        lookup[k] = v

# ── SCAN FILES ────────────────────────────────────────────────────────────────
SKIP = {
    "Eat Clean - Breakfast", "Eat Clean - Snacks", "Eat Clean - Mains",
    "Eat Clean - Additional Items", "Type",
    "100 Carrots",
    "Breakfast - High Protein150g Sunny Side Up Eggs (approx 3)",
}

files = sorted(glob.glob(os.path.join(sel_tmp, "**", "*.xlsx"), recursive=True))
print(f"  {len(files)} selection files found\n")

zero_price  = {}   # dish → [(date, client)]
unrecognised = {}  # dish → [(date, client)]  — starts with digit (gram-based) but not standard

for f in files:
    try:
        df = pd.read_excel(f, header=0)
    except Exception:
        continue

    if "Person" not in df.columns:
        continue

    for _, row in df.iterrows():
        person = str(row.get("Person", "")).strip()
        date   = str(row.get("Meal plan for the day", ""))[:10]

        # Only process dates in the last 30 days
        if date < START_DATE or date > CUTOFF_DATE:
            continue

        for line in str(row.get("Meals", "")).split("\n"):
            line = line.strip()
            if not line or ":" not in line:
                continue
            mt, dish_raw = line.split(":", 1)
            dish = re.sub(r"\s*-\s*\d+\s*$", "", dish_raw.strip().replace("(x)", "")).strip()
            dish = re.sub(r"\s+", " ", dish)

            if not dish or dish in SKIP:
                continue

            # Gram-based (HP/MYO items) — skip
            if re.match(r"^\d+g?\s", dish) or re.match(r"^\d+x?\s+\d+g", dish):
                continue

            entry = (date, person)

            if dish in lookup:
                pass  # priced correctly
            elif dish.lower() in {k.lower() for k in lookup}:
                # Case mismatch
                if dish not in zero_price:
                    zero_price[dish] = []
                zero_price[dish].append(entry)
            else:
                # Completely missing
                if dish not in zero_price:
                    zero_price[dish] = []
                zero_price[dish].append(entry)

# ── REPORT ────────────────────────────────────────────────────────────────────
print("=" * 70)
print(f"DISH PRICE CHECK — last 30 days ({START_DATE} → {CUTOFF_DATE})")
print("=" * 70)

if not zero_price:
    print("\n✅ No zero-price or unrecognised dishes found. All clear.")
else:
    print(f"\n❌ {len(zero_price)} dish(es) with missing prices:\n")
    print(f"  {'Dish':<50} {'Date':<12} {'Client'}")
    print("  " + "-" * 90)
    for dish in sorted(zero_price, key=lambda x: -len(zero_price[x])):
        hits = zero_price[dish]
        # Show first occurrence + count
        first_date, first_client = hits[0]
        total = len(hits)
        print(f"  {dish:<50} {first_date:<12} {first_client}")
        if total > 1:
            for d, c in hits[1:]:
                print(f"  {'':50} {d:<12} {c}")

print("\n" + "=" * 70)
print(f"Period: {START_DATE} → {CUTOFF_DATE} | Price list: {len(lookup)} dishes")
print("=" * 70)
