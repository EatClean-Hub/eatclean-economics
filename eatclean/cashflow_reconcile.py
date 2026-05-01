"""
Reconcile 2026 cash flow (from Wio statement) against the 2026 P&L
(INCOME STATEMENT FST - PL_March 2026.xlsx).

For each P&L line in Jan, Feb, Mar 2026, compare the reported P&L figure
to the cash actually paid/received on the bank statement. The difference
is the accrual: invoices booked but not yet paid, or cash paid ahead of expense.

Reads cashflow.json (produced by cashflow_build.py) and the P&L workbook
from Drive. Emits:
  /tmp/eatclean_cashflow_2026/reconciliation.csv
  /tmp/eatclean_cashflow_2026/reconciliation.json
"""

import sys, os, csv, json
sys.path.insert(0, os.path.dirname(__file__))

from drive_client import get_service, download_file

PL_FILE_ID = "1YP1uAPdq2gnUBwToxLdmKXE1NqurTTfy"  # INCOME STATEMENT FST - PL_March 2026.xlsx
PL_LOCAL = "/tmp/eatclean_cashflow_2026/pl_march_2026.xlsx"
CASHFLOW_JSON = "/tmp/eatclean_cashflow_2026/cashflow.json"

# Mapping: cash flow (group, line) -> P&L row label
# Cash flow aggregates a few things differently than P&L, so map many:one where needed.
# The P&L labels come from the 2026 sheet we read earlier.
CF_TO_PL = {
    # (cf_group, cf_line) -> P&L row label (case-insensitive substring match)
    ("Revenue", "Stripe"):            "Stripe",
    ("Revenue", "Meal Planet"):       "Meal Planet",
    ("Revenue", "Class Pass"):        "Class Pass",
    ("Revenue", "Tabby"):             "Tabby",
    ("Revenue", "Pointspay"):         "Pointspay",
    ("COGS", "COGS Basiligo"):        "COGS BASILIGO",
    ("COGS", "COGS Delivery fee LogX"): "COGS Delivery fee LogX",
    ("COGS", "COGS Packaging"):       "COGS Packaging",
    ("COGS", "COGS Stickers"):        "COGS Stickers",
    ("Salaries", "Salaries and Wages UAE"):    "Salaries and Wages UAE",
    ("Salaries", "Salaries and Wages Remote"): "Salaries and Wages Remote",
    ("Salaries", "Insurance expenses"):        "Insurance expenses",
    ("Salaries", "Visa expenses"):             "Visa expenses",
    ("Salaries", "Gratuities"):                "Gratuities",
    ("Salaries", "Employee benefits"):         "Employee benefits",
    ("Advertising", "Agency Retainer Fee"):    "Agency Retainer Fee",
    ("Advertising", "Google Ads expenses"):    "Google Ads expenses",
    ("Advertising", "Facebook/Instagram Expenses"): "Facebook/Instagram Expenses",
    ("Advertising", "Other advertising"):      "Other advertising",
    ("CRM + Platforms", "Nutribot"):           "Nutribot",
    ("CRM + Platforms", "Shopify + website"):  "Shopify + website",
    ("CRM + Platforms", "Suscription fee"):    "Suscription fee",
    ("CRM + Platforms", "Google Space"):       "Google Space",
    ("CRM + Platforms", "GoDaddy"):            "GoDaddy",
    ("Sales Fees", "Stripe Fees"):             "Stripe Fees",
    ("Sales Fees", "Meal Planet Fees"):        "Meal Planet Fees",
    ("Accounting", "Salary - Accountant"):     "Salary - Accountant",
    ("Accounting", "Audit Fee"):               "Audit Fee",
    ("Finance costs", "Bank Fees and Charges"):"Bank Fees and Charges",
    ("Legal costs", "Rent Expense"):           "Rent Expense",
    ("Legal costs", "Trade license expenses"): "Trade license expenses",
    ("Other expenses", "Other"):               "Other expenses",
}


def load_pl() -> dict:
    """Parse the 2026 income statement and return {pl_label: {month: value}}.
    Locates the header row dynamically since the sheet has leading empty rows.
    """
    if not os.path.exists(PL_LOCAL):
        svc = get_service()
        download_file(svc, PL_FILE_ID, PL_LOCAL)
    import openpyxl
    wb = openpyxl.load_workbook(PL_LOCAL, data_only=True)
    ws = wb["INCOME STATEMENT 2026"]

    # Find the header row (first cell == "PARTICULATRS" or starts with "PART")
    header_row_idx = None
    for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
        if row and isinstance(row[0], str) and row[0].strip().upper().startswith("PART"):
            header_row_idx = ws.max_row  # placeholder — re-get below
            break

    # Re-scan for the row number
    header_row_num = None
    for r in range(1, 25):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.strip().upper().startswith("PART"):
            header_row_num = r
            break
    if header_row_num is None:
        raise RuntimeError("Could not find PARTICULATRS header row in P&L sheet")

    header = [c.value for c in ws[header_row_num]]
    month_cols = {}
    for idx, v in enumerate(header):
        if hasattr(v, "year") and v.year == 2026:
            month_cols[f"2026-{v.month:02d}"] = idx

    data = {}
    for row in ws.iter_rows(min_row=header_row_num + 1, values_only=True):
        label = row[0]
        if not label or not isinstance(label, str):
            continue
        per_month = {}
        for mk, ci in month_cols.items():
            v = row[ci] if ci < len(row) else None
            try:
                per_month[mk] = float(v) if v not in (None, "") else 0.0
            except (TypeError, ValueError):
                per_month[mk] = 0.0
        data[label.strip()] = per_month
    return data


def main():
    with open(CASHFLOW_JSON) as f:
        cashflow = json.load(f)
    pl = load_pl()

    months = ["2026-01", "2026-02", "2026-03"]  # P&L has actual data only for these
    rows = []

    # Cash flow sign is already signed (inflow +, outflow -).
    # P&L reports expenses as positive numbers; revenue as positive.
    # To compare: take abs of cash flow so both are magnitudes.

    for (cf_group, cf_line), pl_label in CF_TO_PL.items():
        cf_per_month = cashflow["buckets"].get(cf_group, {}).get(cf_line, {})
        pl_per_month = pl.get(pl_label, {})
        row = {"cf_group": cf_group, "cf_line": cf_line, "pl_label": pl_label}
        for m in months:
            cf_val = abs(cf_per_month.get(m, 0.0))
            pl_val = abs(pl_per_month.get(m, 0.0))
            row[f"cf_{m}"] = round(cf_val, 2)
            row[f"pl_{m}"] = round(pl_val, 2)
            row[f"delta_{m}"] = round(cf_val - pl_val, 2)  # +ve = cash paid > booked; -ve = accrued
        rows.append(row)

    # Write CSV
    csv_path = "/tmp/eatclean_cashflow_2026/reconciliation.csv"
    with open(csv_path, "w", newline="") as f:
        fields = ["cf_group", "cf_line", "pl_label"]
        for m in months:
            fields += [f"cf_{m}", f"pl_{m}", f"delta_{m}"]
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for r in rows:
            w.writerow(r)

    # JSON
    json_path = "/tmp/eatclean_cashflow_2026/reconciliation.json"
    with open(json_path, "w") as f:
        json.dump(rows, f, indent=2)

    # Print a summary
    print(f"{'Line':45} | " + " | ".join(
        f"{m} CF    P&L    Δ" for m in months
    ))
    print("-" * 140)
    for r in rows:
        tag = r["cf_line"][:43]
        parts = []
        for m in months:
            parts.append(f"{r[f'cf_{m}']:>10,.0f} {r[f'pl_{m}']:>10,.0f} {r[f'delta_{m}']:>+10,.0f}")
        print(f"{tag:45} | " + " | ".join(parts))

    print(f"\nOutputs:\n  {csv_path}\n  {json_path}")


if __name__ == "__main__":
    main()
