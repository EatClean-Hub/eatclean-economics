"""
EatClean Unit Economics Pipeline
Produces the same format as v16 Assessment:
- Per-customer rows with daily kitchen cost columns
- 3 sheets: Regular Customers | HP — High Protein | Valorem — Corporate
Run: python3 run.py
"""

import os, re, glob, tempfile
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import TARGET_MARGIN, DELIVERY_PER_DAY, PACKAGING_PER_DAY, VAT_DIVISOR, CLASSPASS_PRICE
from drive_client import get_service, discover_folders, download_folder_to_temp, upload_file
from parsers import (build_price_lookup, build_gram_lookup,
                     load_all_selections, load_all_invoices, load_orders, _normalize, _normalize_variants)

# ─── STYLES ───────────────────────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", start_color="1F4E79")
DATE_FILL  = PatternFill("solid", start_color="2E75B6")
HP_FILL    = PatternFill("solid", start_color="7030A0")
HP_DATE    = PatternFill("solid", start_color="9966CC")
VAL_FILL   = PatternFill("solid", start_color="833C00")
VAL_DATE   = PatternFill("solid", start_color="C55A11")
GOOD_FILL  = PatternFill("solid", start_color="E2EFDA")
WARN_FILL  = PatternFill("solid", start_color="FCE4D6")
MID_FILL   = PatternFill("solid", start_color="FFF2CC")
GREY_FILL  = PatternFill("solid", start_color="F2F2F2")
COMP_FILL  = PatternFill("solid", start_color="C6EFCE")
PART_FILL  = PatternFill("solid", start_color="FFEB9C")
DEP_FILL   = PatternFill("solid", start_color="FCE4D6")
DAYS_FILL  = PatternFill("solid", start_color="D9E1F2")
DATE_CELL  = PatternFill("solid", start_color="DEEAF1")
HP_CELL    = PatternFill("solid", start_color="EAD1F5")
VAL_CELL   = PatternFill("solid", start_color="FCE4C9")
THIN       = Side(style="thin", color="CCCCCC")
BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CTR        = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT        = Alignment(horizontal="left",   vertical="center")
HDR_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT  = Font(name="Arial", size=10)

FIXED_COLS = ["Customer","Meal Plan","Meals","Paid Days","Actual Days",
              "Delivery Status",
              "Paid ex-VAT (AED)","Deposit (AED)","Net Revenue (AED)",
              "Order Start","Order End","Data Complete","Late Delivery Days",
              "Kitchen Cost (AED)","Delivery Cost (AED)","Packaging Cost (AED)",
              "Total Cost (AED)","Margin (AED)","Margin %"]
AED_COLS   = {"Paid ex-VAT (AED)","Deposit (AED)","Net Revenue (AED)",
              "Kitchen Cost (AED)","Delivery Cost (AED)","Packaging Cost (AED)",
              "Total Cost (AED)","Margin (AED)"}
DATE_START = len(FIXED_COLS) + 1  # auto-calculated from FIXED_COLS


# ══════════════════════════════════════════════════════════════════════════════
# BUILD WIDE TABLE — matches v16 logic exactly
# ══════════════════════════════════════════════════════════════════════════════

def build_wide(orders_df: pd.DataFrame,
               daily_df: pd.DataFrame,
               all_dates: list,
               fixed_revenue: float = None) -> pd.DataFrame:
    """
    Build per-order wide table with daily cost columns.
    orders_df: filtered orders for this group
    daily_df:  all daily selections for this group (person × date × line_cost)
    all_dates: list of date strings ('2026-03-01' etc) for column headers
    """
    # Build daily cost per person from selections
    if not daily_df.empty:
        daily_cost = (daily_df.groupby(["person","date"])["line_cost"]
                      .sum().reset_index()
                      .rename(columns={"line_cost":"day_cost"}))
        daily_cost["person_norm"] = daily_cost["person"].apply(_normalize)
        daily_cost["date"] = pd.to_datetime(daily_cost["date"])
    else:
        daily_cost = pd.DataFrame(columns=["person_norm","date","day_cost"])

    data_start = pd.Timestamp(all_dates[0])  if all_dates else None
    data_end   = pd.Timestamp(all_dates[-1]) if all_dates else None

    rows = []
    for _, order in orders_df.iterrows():
        person_norm = order["name_norm"]
        order_start = order["Order date from"]
        order_end   = order["Order date to"]
        paid_days   = int(order["Delivery period"])
        net_rev     = fixed_revenue if fixed_revenue is not None else float(order["net_revenue"])
        deposit     = float(order["Deposits"])

        # Order date range overlaps data range?
        data_complete = (data_start is not None and
                         order_start >= data_start and
                         order_end   <= data_end)

        # Match person in daily selections — try all name variants (handles reversed names)
        name_variants = _normalize_variants(order["full_name"])
        # Match within order period first
        person_data = daily_cost[
            (daily_cost["person_norm"].isin(name_variants)) &
            (daily_cost["date"] >= order_start) &
            (daily_cost["date"] <= order_end)
        ] if not daily_cost.empty else pd.DataFrame()

        # Late delivery buffer: check selections AFTER order_end
        # Covers: late deliveries, holidays, plan extensions
        # Cap: exclude dates that fall within another order for the same client
        other_order_dates = set()
        for other_idx, other in orders_df.iterrows():
            if other_idx == order.name: continue  # skip current order
            if _normalize(other["full_name"]) in name_variants:
                # Another order for same client — exclude its dates
                o_start = other["Order date from"]
                o_end   = other["Order date to"]
                if not daily_cost.empty:
                    overlap = daily_cost[
                        (daily_cost["date"] >= o_start) &
                        (daily_cost["date"] <= o_end)
                    ]["date"]
                    other_order_dates.update(overlap.tolist())

        late_data = daily_cost[
            (daily_cost["person_norm"].isin(name_variants)) &
            (daily_cost["date"] > order_end) &
            (~daily_cost["date"].isin(other_order_dates))
        ] if not daily_cost.empty else pd.DataFrame()
        if not late_data.empty:
            late_days  = sorted(late_data["date"].dt.strftime("%Y-%m-%d").tolist())
            person_data = pd.concat([person_data, late_data], ignore_index=True)
        else:
            late_days = []

        actual_days   = len(person_data)
        kitchen_cost  = round(person_data["day_cost"].sum(), 2) if not person_data.empty else 0

        # ── STEP 5: Delivery = (actual_days + 1) × 14 — includes final bag collection ──
        delivery_cost = round((actual_days + 1) * DELIVERY_PER_DAY, 2) if actual_days > 0 else 0
        pack_cost     = round(actual_days * PACKAGING_PER_DAY, 2)
        total_cost    = round(kitchen_cost + delivery_cost + pack_cost, 2)

        # ── STEP 5: Delivery status flags ────────────────────────────────────
        if actual_days >= paid_days:
            delivery_status = "✅ Match"
        elif data_end is not None and order_end > data_end:
            delivery_status = "⏳ Ongoing"   # plan not finished yet
        else:
            delivery_status = "🔴 Check"     # plan finished but days missing

        # Coverage ratio — how much of the paid period do we have data for
        coverage = actual_days / paid_days if paid_days > 0 else 0

        if kitchen_cost > 0 and coverage >= 0.8:
            margin_aed = round(net_rev - total_cost, 2)
            margin_pct = round(margin_aed / net_rev * 100, 1) if net_rev > 0 else None
        elif kitchen_cost > 0 and actual_days > 0:
            avg_daily_cost = (kitchen_cost + delivery_cost + pack_cost) / actual_days
            est_total_cost = round(avg_daily_cost * paid_days, 2)
            margin_aed     = round(net_rev - est_total_cost, 2)
            margin_pct     = round(margin_aed / net_rev * 100, 1) if net_rev > 0 else None
        else:
            margin_aed = None
            margin_pct = None

        plan_parts = str(order.get("Order","")).split(",")
        row = {
            "Customer":            order["full_name"],
            "Meal Plan":           plan_parts[0].strip() if plan_parts else "",
            "Meals":               plan_parts[1].strip() if len(plan_parts)>1 else "",
            "Paid Days":           paid_days,
            "Actual Days":         actual_days if actual_days > 0 else None,
            "Delivery Status":     delivery_status,
            "Paid ex-VAT (AED)":   order["paid_ex_vat"],
            "Deposit (AED)":       deposit if deposit > 0 else None,
            "Net Revenue (AED)":   net_rev,
            "Order Start":         order_start.strftime("%d %b %Y"),
            "Order End":           order_end.strftime("%d %b %Y"),
            "Data Complete":       "✅ Yes" if data_complete else "⏳ Partial",
            "Late Delivery Days":  ", ".join(late_days) if late_days else None,
            "Kitchen Cost (AED)":  kitchen_cost if kitchen_cost > 0 else None,
            "Delivery Cost (AED)": delivery_cost if actual_days > 0 else None,
            "Packaging Cost (AED)":pack_cost     if actual_days > 0 else None,
            "Total Cost (AED)":    total_cost    if kitchen_cost > 0 else None,
            "Margin (AED)":        margin_aed,
            "Margin %":            margin_pct,
        }

        # Daily columns — include regular days AND late delivery days
        for d in all_dates:
            d_ts = pd.Timestamp(d)
            if not person_data.empty:
                day_rows = person_data[person_data["date"] == d_ts]
                if not day_rows.empty:
                    cost = round(day_rows["day_cost"].sum(), 2)
                    row[d] = cost if cost > 0 else None
                else:
                    row[d] = None
            else:
                row[d] = None
        rows.append(row)

    return pd.DataFrame(rows)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET BUILDER — v16 style
# ══════════════════════════════════════════════════════════════════════════════

def build_sheet(ws, df_wide, all_dates, title,
                main_fill, d_hdr_fill, d_cell_fill, note=None):

    all_cols = FIXED_COLS + all_dates

    # Row 1 — section headers
    ws.cell(row=1, column=1).value = f"ORDER INFO  |  {title}"
    ws.merge_cells(start_row=1, end_row=1,
                   start_column=1, end_column=len(FIXED_COLS))
    c = ws.cell(row=1, column=1)
    c.font=HDR_FONT; c.fill=main_fill; c.alignment=CTR; c.border=BORDER

    label = "DAILY KITCHEN COST (AED)" + (f"  |  {note}" if note else "")
    ws.cell(row=1, column=DATE_START).value = label
    if all_dates:  # only merge if there are date columns
        ws.merge_cells(start_row=1, end_row=1,
                       start_column=DATE_START,
                       end_column=DATE_START+len(all_dates)-1)
    c = ws.cell(row=1, column=DATE_START)
    c.font=HDR_FONT; c.fill=d_hdr_fill; c.alignment=CTR; c.border=BORDER
    ws.row_dimensions[1].height = 18

    # Row 2 — column headers
    for col, h in enumerate(all_cols, 1):
        c = ws.cell(row=2, column=col)
        c.font = HDR_FONT; c.border = BORDER
        c.fill = d_hdr_fill if col >= DATE_START else main_fill
        c.alignment = CTR
        if col >= DATE_START:
            try:
                c.value = datetime.strptime(h, "%Y-%m-%d").strftime("%d %b")
            except:
                c.value = h
        else:
            c.value = h
    ws.row_dimensions[2].height = 28

    # Data rows
    for _, row_data in df_wide.iterrows():
        ws.append([row_data.get(c) for c in all_cols])
        r = ws.max_row
        mp = row_data.get("Margin %")
        row_fill = (GOOD_FILL if (mp is not None and mp >= TARGET_MARGIN)
                    else MID_FILL  if (mp is not None and mp >= 30)
                    else WARN_FILL if mp is not None
                    else GREY_FILL)
        for col in range(1, len(all_cols)+1):
            c   = ws.cell(row=r, column=col)
            h   = all_cols[col-1]
            c.font = BODY_FONT; c.border = BORDER
            if col < DATE_START:
                if h == "Data Complete":
                    c.fill = COMP_FILL if c.value == "✅ Yes" else PART_FILL
                    c.alignment = CTR
                elif h == "Actual Days":
                    c.fill = DAYS_FILL; c.alignment = CTR
                elif h == "Delivery Status":
                    if c.value == "🔴 Check":
                        c.fill = PatternFill("solid", start_color="FFCCCC")
                        c.font = Font(name="Arial", bold=True, size=10, color="CC0000")
                    elif c.value == "⏳ Ongoing":
                        c.fill = PatternFill("solid", start_color="FFEB9C")
                    else:
                        c.fill = PatternFill("solid", start_color="C6EFCE")
                    c.alignment = CTR
                elif h == "Deposit (AED)" and c.value:
                    c.fill = DEP_FILL; c.alignment = CTR
                else:
                    c.fill = row_fill
                    c.alignment = LFT if h == "Customer" else CTR
                if h in AED_COLS:
                    c.number_format = "#,##0.00"
                if h == "Margin %":
                    c.number_format = '0.0"%"'
            else:
                val = c.value
                c.fill = d_cell_fill if (val and val > 0) else PatternFill()
                c.alignment = CTR
                if val:
                    c.number_format = "#,##0.00"

    # Column widths
    widths = {
        "Customer":26,"Meal Plan":20,"Meals":13,"Paid Days":10,"Actual Days":11,"Delivery Status":14,
        "Paid ex-VAT (AED)":14,"Deposit (AED)":12,"Net Revenue (AED)":15,
        "Order Start":11,"Order End":11,"Data Complete":12,
        "Kitchen Cost (AED)":14,"Delivery Cost (AED)":14,"Packaging Cost (AED)":14,
        "Total Cost (AED)":14,"Margin (AED)":13,"Margin %":10,
    }
    for col, h in enumerate(all_cols, 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(h, 10)

    ws.freeze_panes = "F3"


# ══════════════════════════════════════════════════════════════════════════════
# VIOLATIONS SHEET
# ══════════════════════════════════════════════════════════════════════════════

def build_violations_sheet(ws, df_violations):
    from config import BENCH_MAX
    hdrs = ["Date","Month","Meal Type","Dish Name","Tier",
            "Invoice Price incl VAT","Invoice Price ex-VAT","Max Allowed ex-VAT","Flag"]
    for col, h in enumerate(hdrs, 1):
        c = ws.cell(row=1, column=col)
        c.value=h; c.font=HDR_FONT; c.fill=HDR_FILL
        c.alignment=CTR; c.border=BORDER

    RED    = PatternFill("solid", start_color="FFCCCC")
    ORANGE = PatternFill("solid", start_color="FFE5CC")

    for _, row in df_violations.iterrows():
        r    = ws.max_row + 1
        flag = row.get("bench_flag","")
        fill = RED if flag=="EXCEEDS ALL" else ORANGE
        tier = row.get("tier") or ""
        mt   = row.get("meal_type","")
        maxb = BENCH_MAX.get((mt, tier),"") if tier else ""
        vals = [str(row.get("date",""))[:10], row.get("month",""), mt,
                row.get("dish_name",""), tier,
                row.get("unit_price_incl_vat"), row.get("unit_price_ex_vat"), maxb,
                "🔴 EXCEEDS ALL" if flag=="EXCEEDS ALL" else "🟡 CHECK"]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=col)
            c.value=val; c.font=BODY_FONT; c.fill=fill; c.border=BORDER
            c.alignment=LFT if col==4 else CTR
            if col in [6,7,8]: c.number_format="#,##0.00"

    ws.column_dimensions["D"].width = 44
    for col in [1,2,3,5,6,7,8,9]:
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.freeze_panes = "A2"


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def _build_dishes_ranking(ws, df_sel):
    """Dishes ranked by total qty ordered — all standard clients."""
    ws.cell(row=1, column=1).value = "DISHES RANKING — by total qty ordered (standard plans, Mar–Apr 2026)"
    ws.merge_cells("A1:F1")
    c = ws.cell(row=1, column=1)
    c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CTR; c.border = BORDER
    ws.row_dimensions[1].height = 20

    headers = ["Rank", "Dish Name", "Total Qty", "Days Ordered",
               "Avg Qty/Day", "Total Cost (AED)"]
    for col, h in enumerate(headers, 1):
        c2 = ws.cell(row=2, column=col)
        c2.value = h; c2.font = HDR_FONT; c2.fill = HDR_FILL
        c2.alignment = CTR; c2.border = BORDER

    if df_sel.empty:
        return

    reg = df_sel[df_sel["group"].isin(["regular", "valorem"])]
    ranked = (
        reg.groupby("dish_name")
        .agg(total_qty=("qty","sum"),
             days_ordered=("date","nunique"),
             total_cost=("line_cost","sum"))
        .reset_index()
        .sort_values("total_qty", ascending=False)
        .reset_index(drop=True)
    )
    ranked["avg_per_day"] = (ranked["total_qty"] / ranked["days_ordered"]).round(1)

    WHITE = PatternFill("solid", start_color="FFFFFF")
    GREY  = PatternFill("solid", start_color="F2F2F2")
    for i, (_, row) in enumerate(ranked.iterrows(), 1):
        r = ws.max_row + 1
        fill = WHITE if i % 2 else GREY
        vals = [i, row["dish_name"], int(row["total_qty"]),
                int(row["days_ordered"]), row["avg_per_day"],
                round(row["total_cost"], 2)]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=col)
            c.value = val; c.font = BODY_FONT; c.fill = fill; c.border = BORDER
            c.alignment = LFT if col == 2 else CTR
            if col == 6: c.number_format = "#,##0.00"

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 16
    ws.freeze_panes = "A3"


def _build_reconciliation(ws, df_inv, df_sel):
    """Invoice vs selections — qty and value comparison per day per dish."""
    ws.cell(row=1, column=1).value = "RECONCILIATION — Invoice vs Daily Selections (BF/MN/SN only, Mar–Apr 2026)"
    ws.merge_cells("A1:H1")
    c = ws.cell(row=1, column=1)
    c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CTR; c.border = BORDER
    ws.row_dimensions[1].height = 20

    headers = ["Date", "Meal Type", "Dish Name",
               "Invoice Qty", "Selection Qty", "Diff (Inv-Sel)",
               "Invoice Cost (AED)", "Flag"]
    for col, h in enumerate(headers, 1):
        c2 = ws.cell(row=2, column=col)
        c2.value = h; c2.font = HDR_FONT; c2.fill = HDR_FILL
        c2.alignment = CTR; c2.border = BORDER

    if df_inv.empty or df_sel.empty:
        return

    # Invoice: aggregate by date × dish (exclude ADD/RT already filtered)
    inv_agg = (df_inv.groupby(["date","meal_type","dish_name"])
               .agg(inv_qty=("qty","sum"),
                    inv_cost=("line_total_ex_vat","sum"))
               .reset_index())

    # Selections: aggregate by date × dish (standard + valorem only)
    reg_sel = df_sel[df_sel["group"].isin(["regular","valorem"])].copy()
    sel_agg = (reg_sel.groupby(["date","dish_name"])
               .agg(sel_qty=("qty","sum"))
               .reset_index())
    sel_agg["date"] = pd.to_datetime(sel_agg["date"])

    # Only dates in BOTH
    inv_agg["date"] = pd.to_datetime(inv_agg["date"])
    common = set(inv_agg["date"].dt.date) & set(sel_agg["date"].dt.date)
    inv_filt = inv_agg[inv_agg["date"].dt.date.isin(common)]
    sel_filt = sel_agg[sel_agg["date"].dt.date.isin(common)]

    merged = pd.merge(inv_filt, sel_filt,
                      on=["date","dish_name"], how="outer")
    merged = merged.fillna(0)
    merged["diff"] = merged["inv_qty"] - merged["sel_qty"]
    merged = merged.sort_values(["date","meal_type","diff"], ascending=[True,True,True])

    RED    = PatternFill("solid", start_color="FFCCCC")
    ORANGE = PatternFill("solid", start_color="FFE5CC")
    GREEN  = PatternFill("solid", start_color="E2EFDA")
    WHITE  = PatternFill("solid", start_color="FFFFFF")

    for _, row in merged.iterrows():
        r = ws.max_row + 1
        diff = row["diff"]
        flag = ("✅ OK" if diff == 0
                else "🔴 OVER-INVOICED" if diff > 0
                else "🟡 UNDER-INVOICED")
        fill = (GREEN if diff == 0 else RED if diff > 0 else ORANGE)
        vals = [str(row["date"])[:10],
                row.get("meal_type",""),
                row["dish_name"],
                int(row["inv_qty"]),
                int(row["sel_qty"]),
                int(diff),
                round(row.get("inv_cost",0), 2),
                flag]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=col)
            c.value = val; c.font = BODY_FONT; c.fill = fill; c.border = BORDER
            c.alignment = LFT if col == 3 else CTR
            if col == 7: c.number_format = "#,##0.00"

    # Summary row
    r = ws.max_row + 2
    ws.cell(row=r, column=1).value = "SUMMARY"
    ws.cell(row=r, column=1).font = Font(name="Arial", bold=True, size=10)
    over  = (merged["diff"] > 0).sum()
    under = (merged["diff"] < 0).sum()
    ok    = (merged["diff"] == 0).sum()
    ws.cell(row=r, column=2).value = f"✅ {ok} matched | 🔴 {over} over-invoiced | 🟡 {under} under-invoiced"

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 44
    ws.column_dimensions["D"].width = 13
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 18
    ws.freeze_panes = "A3"


def main():
    print("=" * 60)
    print("EatClean Unit Economics Pipeline")
    print(f"Run: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 60)

    # ── Connect to Drive ─────────────────────────────────────────────────────
    print("\nConnecting to Google Drive...")
    try:
        service = get_service()
        folders = discover_folders(service)
        print(f"  ✅ Connected — {list(folders.keys())}")
    except Exception as e:
        print(f"  ❌ {e}"); return

    with tempfile.TemporaryDirectory() as tmp:

        # ── Download ──────────────────────────────────────────────────────────
        print("\nDownloading files...")
        sel_tmp = os.path.join(tmp, "daily_selections")
        inv_tmp = os.path.join(tmp, "invoices")
        ref_tmp = os.path.join(tmp, "reference")
        out_tmp = os.path.join(tmp, "outputs")
        for d in [sel_tmp, inv_tmp, ref_tmp, out_tmp]:
            os.makedirs(d, exist_ok=True)

        if "daily_selections" in folders:
            import shutil
            cache_dir = os.path.join(os.path.dirname(__file__), ".selections_cache")
            os.makedirs(cache_dir, exist_ok=True)

            # Get list of remote files with their IDs
            from drive_client import list_folder, list_subfolders, download_file
            sel_folder_id = folders["daily_selections"]
            remote_files = []
            try:
                for sf in list_subfolders(service, sel_folder_id):
                    for f in list_folder(service, sf["id"]):
                        if f["name"].endswith(".xlsx"):
                            remote_files.append((f["id"], f["name"]))
                for f in list_folder(service, sel_folder_id):
                    if f["name"].endswith(".xlsx"):
                        remote_files.append((f["id"], f["name"]))
            except Exception as e:
                print(f"  [WARN] Could not list remote files: {e}")

            new_files = cached_files = 0
            if remote_files:
                for fid, fname in remote_files:
                    cached_path = os.path.join(cache_dir, fname)
                    dest_path   = os.path.join(sel_tmp, fname)
                    if os.path.exists(cached_path):
                        shutil.copy2(cached_path, dest_path)
                        cached_files += 1
                    else:
                        try:
                            download_file(service, fid, dest_path)
                            shutil.copy2(dest_path, cached_path)
                            new_files += 1
                        except Exception:
                            pass
            else:
                # Fallback: download everything
                download_folder_to_temp(service, folders["daily_selections"], sel_tmp)

            # Put files into date-based subfolders matching Drive structure
            for f in glob.glob(os.path.join(sel_tmp, "*.xlsx")):
                fname = os.path.basename(f)
                m = re.search(r"(\d{4}-\d{2})", fname)
                if m:
                    month_dir = os.path.join(sel_tmp, m.group(1))
                    os.makedirs(month_dir, exist_ok=True)
                    import shutil as _sh
                    _sh.move(f, os.path.join(month_dir, fname))

            n = sum(len(glob.glob(os.path.join(sel_tmp, "**", "*.xlsx"), recursive=True)) for _ in [1])
            print(f"  Selections: {n} files ({cached_files} cached, {new_files} new)")
        if "invoices" in folders:
            download_folder_to_temp(service, folders["invoices"], inv_tmp)
            n = len(glob.glob(os.path.join(inv_tmp,"**","*.xlsx"), recursive=True))
            print(f"  Invoices: {n} files")
        if "reference" in folders:
            download_folder_to_temp(service, folders["reference"], ref_tmp)

        # Override paths
        import config
        config.PATHS.update({"daily_selections":sel_tmp,"invoices":inv_tmp,
                              "reference":ref_tmp,"outputs":out_tmp})

        # ── Price lookup — invoice file is the ONLY source ───────────────────
        print("\nBuilding price lookup...")
        inv_files  = glob.glob(os.path.join(inv_tmp, "**", "*.xlsx"), recursive=True)
        # Invoice = Eat Clean Sales file; dishes_list = separate file
        price_path      = next((f for f in inv_files
                                if "eat clean sales" in os.path.basename(f).lower()), None)
        dishes_list_path = next((f for f in inv_files
                                 if "dishes_list" in os.path.basename(f).lower()), None)
        if not price_path:
            # Fallback: first file that is NOT dishes_list
            price_path = next((f for f in inv_files
                               if "dishes_list" not in os.path.basename(f).lower()), None)
        if not price_path:
            print("  ❌ No invoice file found in 02 Invoices — cannot build price lookup")
            return
        print(f"  Invoice: {os.path.basename(price_path)}")
        if dishes_list_path:
            print(f"  Dishes list: {os.path.basename(dishes_list_path)}")
        lookup      = build_price_lookup(price_path, dishes_list_path)
        gram_lookup = build_gram_lookup(price_path)

        # ── Orders ────────────────────────────────────────────────────────────
        print("\nLoading orders...")
        try:
            df_orders = load_orders()
            # Filter to 2026 only
            df_orders = df_orders[df_orders["Order date from"] >= "2026-01-01"].copy()
            hp_norms  = set(df_orders[df_orders["is_hp"]]["name_norm"].tolist())
            print(f"  {len(df_orders)} orders (2026) | {len(hp_norms)} HP clients")
        except Exception as e:
            print(f"  [WARN] {e}"); df_orders = pd.DataFrame(); hp_norms = set()

        # ── Daily selections ──────────────────────────────────────────────────
        print("\nLoading daily selections...")
        df_sel = load_all_selections(lookup, gram_lookup, hp_norms)
        if not df_sel.empty:
            all_dates = sorted(df_sel["date"].dt.strftime("%Y-%m-%d").unique().tolist())
            print(f"  {len(df_sel)} lines | {len(all_dates)} days | "
                  f"{df_sel['person'].nunique()} people")
        else:
            all_dates = []
            print("  [WARN] No selection data")

        # ── Invoices (for violations sheet) ───────────────────────────────────
        print("\nLoading invoices...")
        df_inv = load_all_invoices()
        if not df_inv.empty:
            print(f"  {len(df_inv)} lines | AED {df_inv['line_total_ex_vat'].sum():,.2f}")

        # ── Split selections by group ─────────────────────────────────────────
        sel_reg  = df_sel[df_sel["group"]=="regular"]   if not df_sel.empty else pd.DataFrame()
        sel_hp   = df_sel[df_sel["group"]=="hp"]        if not df_sel.empty else pd.DataFrame()
        sel_val  = df_sel[df_sel["group"]=="valorem"]   if not df_sel.empty else pd.DataFrame()
        sel_cp   = df_sel[df_sel["group"]=="classpass"] if not df_sel.empty else pd.DataFrame()
        # Complimentary — uses regular selections (comp clients appear as regular in daily files)
        sel_comp = df_sel[df_sel["group"].isin(["regular", "complimentary"])] if not df_sel.empty else pd.DataFrame()

        # ── Split orders by group ─────────────────────────────────────────────
        if not df_orders.empty:
            ord_reg  = df_orders[~df_orders["is_hp"] & ~df_orders["is_val"] &
                                  ~df_orders["is_cp"] & ~df_orders["is_comp"]]
            ord_hp   = df_orders[df_orders["is_hp"]  & ~df_orders["is_val"]]
            ord_val  = df_orders[df_orders["is_val"]]
            ord_cp   = df_orders[df_orders["is_cp"]]
            ord_comp = df_orders[df_orders["is_comp"]]
        else:
            ord_reg = ord_hp = ord_val = ord_cp = ord_comp = pd.DataFrame()

        # ── Pre-scan for late delivery dates before building wide tables ─────
        # Find any selection dates beyond order_end for any client
        # so all_dates includes those columns before build_wide runs
        if not df_sel.empty and not df_orders.empty:
            sel_dates = set(df_sel["date"].dt.normalize().unique())
            extra_dates = set()
            for _, order in df_orders.iterrows():
                order_end = order["Order date to"]
                name_vars = _normalize_variants(order["full_name"])
                late = df_sel[
                    (df_sel["date"] > order_end) &
                    (df_sel["person"].apply(_normalize).isin(name_vars))
                ]
                for d in late["date"].dt.strftime("%Y-%m-%d").unique():
                    if d not in all_dates:
                        extra_dates.add(d)
            if extra_dates:
                all_dates = sorted(set(all_dates) | extra_dates)
                print(f"  ⏰ {len(extra_dates)} late delivery date(s) pre-scanned: {sorted(extra_dates)}")

        # ── Build wide tables ──────────────────────────────────────────────────
        print("\nBuilding assessment tables...")
        df_wide_reg  = build_wide(ord_reg,  sel_reg,  all_dates) if not ord_reg.empty  else pd.DataFrame()
        df_wide_hp   = build_wide(ord_hp,   df_sel,   all_dates) if not ord_hp.empty   else pd.DataFrame()  # uses all selections for cross-group name matching
        df_wide_val  = build_wide(ord_val,  df_sel,   all_dates) if not ord_val.empty  else pd.DataFrame()  # uses all selections for cross-group name matching
        df_wide_cp   = build_wide(ord_cp,   sel_cp,   all_dates, fixed_revenue=CLASSPASS_PRICE) if not ord_cp.empty   else pd.DataFrame()
        df_wide_comp = build_wide(ord_comp, sel_comp, all_dates) if not ord_comp.empty else pd.DataFrame()

        # Print margin summary
        for label, df in [("Regular", df_wide_reg), ("HP", df_wide_hp),
                          ("Valorem", df_wide_val), ("ClassPass", df_wide_cp),
                          ("Complimentary", df_wide_comp)]:
            if not df.empty and "Margin %" in df.columns:
                with_data = df[df["Margin %"].notna()]
                below = with_data[with_data["Margin %"] < TARGET_MARGIN]
                print(f"  {label}: {len(df)} orders | "
                      f"{len(with_data)} with margin data | "
                      f"🔴 {len(below)} below {TARGET_MARGIN}%")

        # ── Violations ────────────────────────────────────────────────────────
        df_viol = df_inv[df_inv["bench_flag"].isin(["EXCEEDS ALL","CHECK"])].copy() \
                  if not df_inv.empty else pd.DataFrame()
        if not df_viol.empty:
            exceed = (df_viol["bench_flag"]=="EXCEEDS ALL").sum()
            print(f"  🔴 {exceed} dish price violations (EXCEEDS ALL)")

        # ── ZERO-PRICE ALERT ─────────────────────────────────────────────────
        df_zero = pd.DataFrame()
        if not df_sel.empty and "zero_price" in df_sel.columns:
            df_zero = (df_sel[df_sel["zero_price"] == True]
                       .groupby("dish_name")
                       .agg(
                           times_ordered=("qty", "sum"),
                           days=("date", "nunique"),
                           clients=("person", lambda x: ", ".join(sorted(set(x))))
                       )
                       .reset_index()
                       .sort_values("times_ordered", ascending=False))

        if not df_zero.empty:
            print("\n" + "!"*60)
            print("⚠️  ZERO-PRICE DISHES DETECTED — MARGINS INCORRECT")
            print("    FOR AFFECTED CUSTOMERS")
            print("!"*60)
            for _, row in df_zero.iterrows():
                print(f"  ❌ \'{row['dish_name']}\'"
                      f" — ordered {row['times_ordered']}x"
                      f" across {row['days']} days")
                print(f"     Clients: {row['clients'][:80]}")
            print("!"*60)
            print("  ACTION: Add missing prices to config.py MANUAL_PRICES")
            print("          then re-run.")
            print("!"*60 + "\n")
        else:
            print("  ✅ All dishes priced — no zero-price issues")

        # ── Build Excel workbook ───────────────────────────────────────────────
        print("\nBuilding workbook...")
        wb = Workbook()

        # Sheet 1: Regular Customers
        ws1 = wb.active; ws1.title = "Regular Customers"
        if not df_wide_reg.empty:
            build_sheet(ws1, df_wide_reg, all_dates,
                        "REGULAR CUSTOMERS", HDR_FILL, DATE_FILL, DATE_CELL)
        else:
            ws1["A1"] = "No regular customer data for 2026"

        # Sheet 2: HP — High Protein
        ws2 = wb.create_sheet("HP — High Protein")
        if not df_wide_hp.empty:
            build_sheet(ws2, df_wide_hp, all_dates,
                        "HP — HIGH PROTEIN", HP_FILL, HP_DATE, HP_CELL)
        else:
            ws2["A1"] = "No HP customer data for 2026"

        # Sheet 3: Valorem — Corporate
        ws3 = wb.create_sheet("Valorem — Corporate")
        if not df_wide_val.empty:
            build_sheet(ws3, df_wide_val, all_dates,
                        "VALOREM — CORPORATE", VAL_FILL, VAL_DATE, VAL_CELL)
        else:
            ws3["A1"] = "No Valorem data for 2026"

        # Sheet 4: ClassPass — cost-only rows, no revenue
        CP_FILL      = PatternFill("solid", start_color="4BACC6")
        CP_DATE_FILL = PatternFill("solid", start_color="9DC3E6")
        CP_CELL_FILL = PatternFill("solid", start_color="DEEAF1")
        ws_cp = wb.create_sheet("ClassPass")
        if not df_wide_cp.empty:
            build_sheet(ws_cp, df_wide_cp, all_dates,
                        "CLASSPASS — Fixed Revenue AED 356.25/order", CP_FILL, CP_DATE_FILL, CP_CELL_FILL,
                        note="Fixed revenue AED 356.25 per order (ClassPass settlement)")
        else:
            ws_cp["A1"] = "No ClassPass data for this period"

        # Sheet 5: Complimentary & Influencers — cost-only rows
        COMP_HDR_FILL  = PatternFill("solid", start_color="7030A0")
        COMP_DATE_FILL = PatternFill("solid", start_color="9966CC")
        COMP_CELL_FILL = PatternFill("solid", start_color="EAD1F5")
        ws_comp = wb.create_sheet("Complimentary & Influencers")
        if not df_wide_comp.empty:
            build_sheet(ws_comp, df_wide_comp, all_dates,
                        "COMPLIMENTARY & INFLUENCERS — COST ONLY", COMP_HDR_FILL, COMP_DATE_FILL, COMP_CELL_FILL,
                        note="No revenue | Kitchen cost tracking only")
        else:
            ws_comp["A1"] = "No Complimentary data for this period"

        # Sheet 4: Dishes Ranking
        ws4 = wb.create_sheet("Dishes Ranking")
        _build_dishes_ranking(ws4, df_sel)

        # Sheet 5: Reconciliation — Invoice vs Selections
        ws5 = wb.create_sheet("Reconciliation")
        _build_reconciliation(ws5, df_inv, df_sel)

        # Sheet 6: Dish Violations
        ws6 = wb.create_sheet("Dish Violations")
        ws6["A1"].value = "DISH PRICE VIOLATIONS"
        ws6["A1"].font = HDR_FONT; ws6["A1"].fill = HDR_FILL
        if not df_viol.empty:
            build_violations_sheet(ws6, df_viol)

        # Sheet 7: ⚠️ Zero Price Alert — always add if any zero-price dishes found
        if not df_zero.empty:
            ws_zero = wb.create_sheet("⚠️ Zero Price Alert")
            ws_zero.sheet_properties.tabColor = "FF0000"
            ws_zero.merge_cells("A1:D1")
            c = ws_zero["A1"]
            c.value = "⚠️ ZERO-PRICE DISHES — MARGINS INCORRECT FOR AFFECTED CUSTOMERS"
            c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
            c.fill = PatternFill("solid", start_color="CC0000")
            c.alignment = CTR; c.border = BORDER
            ws_zero.row_dimensions[1].height = 22
            ALERT_FILL = PatternFill("solid", start_color="FFCCCC")
            for col, h in enumerate(["Dish Name","Times Ordered","Days Affected","Clients Affected"], 1):
                c = ws_zero.cell(row=2, column=col)
                c.value = h; c.font = HDR_FONT
                c.fill = PatternFill("solid", start_color="CC0000")
                c.alignment = CTR; c.border = BORDER
            for _, row in df_zero.iterrows():
                r = ws_zero.max_row + 1
                for col, val in enumerate([
                    row["dish_name"], int(row["times_ordered"]),
                    int(row["days"]), row["clients"]
                ], 1):
                    c = ws_zero.cell(row=r, column=col)
                    c.value = val; c.font = BODY_FONT
                    c.fill = ALERT_FILL; c.border = BORDER
                    c.alignment = LFT if col in [1,4] else CTR
            ws_zero.column_dimensions["A"].width = 50
            ws_zero.column_dimensions["B"].width = 14
            ws_zero.column_dimensions["C"].width = 14
            ws_zero.column_dimensions["D"].width = 60
            ws_zero.freeze_panes = "A3"

        # ── Save & upload ─────────────────────────────────────────────────────
        timestamp    = datetime.now().strftime("%Y%m%d_%H%M")
        report_name  = f"EatClean_UnitEconomics_{timestamp}.xlsx"

        # Primary: save to Google Drive local folder
        drive_output = os.path.expanduser(
            "~/Library/CloudStorage/GoogleDrive-admin@eatcleanme.com"
            "/My Drive/02 Automation /04 Outputs"
        )
        if os.path.exists(drive_output):
            local_report = os.path.join(drive_output, report_name)
            wb.save(local_report)
            print(f"\n✅ Report saved to Drive 04 Outputs: {report_name}")
        else:
            # Fallback: Desktop
            local_report = os.path.join(os.path.expanduser("~/Desktop"), report_name)
            wb.save(local_report)
            print(f"\n✅ Report saved to Desktop: {report_name}")
            print(f"   (Drive folder not accessible — using Desktop as fallback)")

    print("\n" + "="*60)
    print("Done.")
    print("="*60)


if __name__ == "__main__":
    main()
