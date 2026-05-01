import os
import re
import glob
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook

from config import (
    PATHS, INVOICE_SHEETS, INVOICE_WEEK_COLS,
    VAT_DIVISOR, BENCH_MIN, BENCH_MAX, MANUAL_PRICES
)

# ─── CONSTANTS ────────────────────────────────────────────────────────────────
TYPE_MAP     = {"BF": "Breakfast", "SN": "Snack", "MN": "Mains"}
SKIP_TYPES   = {"ADD", "RT"}
SKIP_DISHES  = {
    "Eat Clean - Breakfast", "Eat Clean - Snacks", "Eat Clean - Mains",
    "Eat Clean - Additional Items", "Type",
    # HP parsing artifacts — malformed lines in standard files
    "100 Carrots",
    "Breakfast - High Protein150g Sunny Side Up Eggs (approx 3)",
}
COMP_KW      = ["reduction", "compensation", "details of"]
COLS_10      = ["customer_id", "person", "address", "meal_plan", "calories",
                "meals", "exclusions", "order_date", "order_period", "nominal_price"]
COLS_9       = ["customer_id", "person", "address", "meal_plan", "calories",
                "meals", "exclusions", "order_date", "nominal_price"]


# ══════════════════════════════════════════════════════════════════════════════
# PRICE LOOKUP
# ══════════════════════════════════════════════════════════════════════════════

def build_price_lookup(price_list_path: str) -> dict:
    """
    Load dish prices from the invoice Excel file (Price List sheet).
    This is the ONLY price source — no fallback files.
    Column A = dish name, Column B = price incl. VAT (divided by 1.05 → ex-VAT).
    MANUAL_PRICES in config.py patches known name mismatches on top.
    """
    lookup = {}

    if not price_list_path or not os.path.exists(price_list_path):
        print("  [WARN] No invoice file found — only MANUAL_PRICES will be used")
    else:
        try:
            wb = load_workbook(price_list_path, read_only=True, data_only=True)
            if "Price List" not in wb.sheetnames:
                print(f"  [WARN] No 'Price List' sheet found in invoice file")
                print(f"         Sheets available: {wb.sheetnames}")
            else:
                ws = wb["Price List"]
                count = 0
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i == 0: continue  # skip header
                    if not row[0] or not row[1]: continue
                    try:
                        name  = str(row[0]).strip()
                        price = float(row[1]) / VAT_DIVISOR
                        if name and price > 0:
                            lookup[name] = round(price, 4)
                            count += 1
                    except (ValueError, TypeError):
                        continue
                print(f"  {count} dishes loaded from invoice Price List")
        except Exception as e:
            print(f"  [WARN] Could not read invoice price list: {e}")

    # Apply MANUAL_PRICES patches on top (case fixes + missing dishes)
    patched = 0
    for k, v in MANUAL_PRICES.items():
        if k not in lookup:
            lookup[k] = v
            patched += 1
    if patched:
        print(f"  {patched} additional entries from MANUAL_PRICES patches")

    return lookup

    try:
        wb = load_workbook(price_list_path, read_only=True, data_only=True)
        sheets = wb.sheetnames

        # ── Invoice format: 'Price List' sheet, col B = Price incl VAT ──────
        if "Price List" in sheets:
            ws = wb["Price List"]
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0: continue  # skip header
                if not row[0] or not row[1]: continue
                try:
                    name  = str(row[0]).strip()
                    price = float(row[1]) / VAT_DIVISOR  # incl → ex VAT
                    if name and price > 0:
                        lookup[name] = round(price, 4)
                except (ValueError, TypeError):
                    continue

        # ── Original price list format: 'Sheet1' = ex-VAT ────────────────────
        if "Sheet1" in sheets:
            ws = wb["Sheet1"]
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0: continue  # skip header
                if not row[0] or not row[1]: continue
                try:
                    name  = str(row[0]).strip()
                    price = float(row[1])
                    if name and price > 0 and name not in lookup:
                        lookup[name] = round(price, 4)
                except (ValueError, TypeError):
                    continue

        # ── New Items sheet ───────────────────────────────────────────────────
        if "New Items" in sheets:
            ws = wb["New Items"]
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0: continue
                if not row[1] or not row[2]: continue
                try:
                    name  = str(row[1]).strip()
                    price = float(row[2])
                    if name and price > 0 and name not in lookup:
                        lookup[name] = round(price, 4)
                except (ValueError, TypeError):
                    continue

    except Exception as e:
        print(f"  [WARN] Price lookup error: {e}")

    # Manual overrides always applied
    for k, v in MANUAL_PRICES.items():
        if k not in lookup:
            lookup[k] = v

    # Build lowercase index for case-insensitive fallback
    lookup["__lower__"] = {k.lower(): v for k, v in lookup.items()
                           if not k.startswith("__")}
    return lookup


def build_gram_lookup(kg_price_list_path: str) -> dict:
    """Load KG-based prices. Returns {ingredient_name_lower: price_per_gram}."""
    gram_lookup = {}
    try:
        df = pd.read_excel(kg_price_list_path, sheet_name="Sheet1", header=1)
        df.columns = ["dish_name", "price", "retail"]
        df = df[["dish_name", "price"]].dropna(subset=["dish_name", "price"])
        df["price"] = pd.to_numeric(df["price"], errors="coerce")
        df["dish_name"] = df["dish_name"].astype(str).str.strip()

        for _, row in df.dropna(subset=["price"]).iterrows():
            name = str(row["dish_name"]).strip()
            if "(KG)" not in name.upper():
                continue
            clean = re.sub(r"\s*\(KG\)\s*$", "", name, flags=re.IGNORECASE).strip()
            clean = re.sub(r"^(PROTEIN|CARBS|SAUCE|VEG)\s*-\s*", "", clean,
                           flags=re.IGNORECASE).strip()
            gram_lookup[clean.lower()] = row["price"] / 1000.0
    except Exception:
        pass
    return gram_lookup


def get_price(dish_name: str, lookup: dict, gram_lookup: dict) -> float:
    """Return ex-VAT price for a dish. Handles fixed and gram-based entries."""
    # 1. Exact match
    if dish_name in lookup:
        return lookup[dish_name]
    # 2. Case-insensitive match via pre-built index
    lower_index = lookup.get("__lower__", {})
    if dish_name.lower() in lower_index:
        return lower_index[dish_name.lower()]
    # 3. Gram-based
    m = re.match(r"^(\d+(?:\.\d+)?)g\s+(.+)$", dish_name, re.IGNORECASE)
    if m:
        weight_g = float(m.group(1))
        ing = m.group(2).strip().lower()
        ppg = gram_lookup.get(ing, 0)
        if ppg == 0:
            for key in gram_lookup:
                if key in ing or ing in key:
                    ppg = gram_lookup[key]
                    break
        if ppg > 0:
            return round(weight_g * ppg, 4)
    return 0.0


# ══════════════════════════════════════════════════════════════════════════════
# DAILY SELECTIONS PARSER
# ══════════════════════════════════════════════════════════════════════════════

def _parse_meal_lines(meals_raw: str, date_str: str,
                      lookup: dict, gram_lookup: dict) -> list:
    """Parse a meal string into list of dicts with dish/qty/cost."""
    records = []
    for line in meals_raw.split("\n"):
        line = line.strip()
        if not line or ":" not in line:
            continue
        if line.upper().startswith("FOOD TO SERVE ONLY"):
            continue
        _, dish_raw = line.split(":", 1)
        dish_raw = dish_raw.strip().replace("(x)", "").strip()
        mult_match = re.match(r"^(\d+)x\s+", dish_raw)
        mult = int(mult_match.group(1)) if mult_match else 1
        if mult_match:
            dish_raw = dish_raw[mult_match.end():]
        dish_name = re.sub(r"\s*-\s*\d+\s*$", "", dish_raw).strip()
        dish_name = re.sub(r"\s+", " ", dish_name)
        if not dish_name or dish_name == "nan":
            continue
        price = get_price(dish_name, lookup, gram_lookup)
        is_gram = bool(re.match(r"^\d+(?:\.\d+)?g\s", dish_name))
        records.append({
            "date":       date_str,
            "dish_name":  dish_name,
            "qty":        mult,
            "unit_cost":  price,
            "line_cost":  round(price * mult, 4),
            "zero_price": (price == 0 and not is_gram),
        })
    return records


def parse_standard_file(filepath: str, lookup: dict, gram_lookup: dict,
                        hp_norms: set) -> dict:
    """Parse standard Nutribot format: ordered-meal-plans-YYYY-MM-DD.xlsx"""
    df = pd.read_excel(filepath, header=0)
    result = {"regular": [], "hp": [], "valorem": [], "classpass": []}

    for _, row in df.iterrows():
        person    = str(row.get("Person", "")).strip()
        date_raw  = str(row.get("Meal plan for the day", "")).strip()
        meals_raw = str(row.get("Meals", "")).strip()
        if str(row.get("Test order", "")).strip().upper() == "YES":
            continue
        date_str = date_raw[:10]
        grp = _route(person, hp_norms)
        lines = _parse_meal_lines(meals_raw, date_str, lookup, gram_lookup)
        for rec in lines:
            rec["person"] = person
        result[grp].extend(lines)
    return result


def parse_report_file(filepath: str, date_str: str,
                      lookup: dict, gram_lookup: dict, hp_norms: set) -> dict:
    """Parse report format: Ordered_Meal_Plans_Report_DD-MM-YYYY.xlsx"""
    sheets = pd.ExcelFile(filepath).sheet_names
    frames = []
    for s in sheets:
        df_s = pd.read_excel(filepath, sheet_name=s, header=0)
        df_s.columns = COLS_10 if len(df_s.columns) == 10 else COLS_9
        frames.append(df_s)
    df = pd.concat(frames, ignore_index=True)

    result = {"regular": [], "hp": [], "valorem": [], "classpass": []}
    for _, row in df.iterrows():
        person    = str(row["person"]).strip()
        meals_raw = str(row["meals"]).strip()
        grp = _route(person, hp_norms)
        lines = _parse_meal_lines(meals_raw, date_str, lookup, gram_lookup)
        for rec in lines:
            rec["person"] = person
        result[grp].extend(lines)
    return result


def _normalize(name: str) -> str:
    name = re.sub(r"^(HP|MP)\s+", "", name.strip(), flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", name).lower().strip()


def _route(person: str, hp_norms: set) -> str:
    """Route to regular / hp / valorem / classpass."""
    upper = person.strip().upper()
    if any(kw in upper for kw in ["VALOREM"]):
        return "valorem"
    if upper.startswith("HP ") or _normalize(person) in hp_norms:
        return "hp"
    if upper.endswith(" CP"):
        return "classpass"
    return "regular"


def load_all_selections(lookup: dict, gram_lookup: dict,
                        hp_norms: set) -> pd.DataFrame:
    """
    Scan all monthly subfolders in 01 Daily_selections.
    De-duplicates: only one file processed per calendar date.
    Returns flat DataFrame: person | date | dish_name | qty | unit_cost | line_cost | group
    """
    base = PATHS["daily_selections"]
    all_records = []
    processed_dates = {}  # date_str → filename (first file wins)

    for month_folder in sorted(os.listdir(base)):
        folder_path = os.path.join(base, month_folder)
        if not os.path.isdir(folder_path):
            continue

        for filepath in sorted(glob.glob(os.path.join(folder_path, "*.xlsx"))):
            fname = os.path.basename(filepath)

            # Standard format: ordered-meal-plans-YYYY-MM-DD.xlsx
            if re.match(r"ordered-meal-plans-\d{4}-\d{2}-\d{2}\.xlsx", fname):
                m = re.search(r"(\d{4}-\d{2}-\d{2})", fname)
                date_str = m.group(1) if m else None
                if date_str and date_str in processed_dates:
                    continue  # skip duplicate date
                recs = parse_standard_file(filepath, lookup, gram_lookup, hp_norms)
                if date_str:
                    processed_dates[date_str] = fname

            # Report format: Ordered_Meal_Plans_Report_DD-MM-YYYY.xlsx
            elif re.match(r"Ordered_Meal_Plans_Report_\d{2}-\d{2}-\d{4}.*\.xlsx", fname, re.IGNORECASE):
                m = re.search(r"(\d{2})-(\d{2})-(\d{4})", fname)
                if m:
                    date_str = f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
                    if date_str in processed_dates:
                        continue  # skip duplicate date
                    recs = parse_report_file(filepath, date_str, lookup,
                                             gram_lookup, hp_norms)
                    processed_dates[date_str] = fname
                else:
                    continue
            else:
                continue

            for grp, records in recs.items():
                for rec in records:
                    rec["group"] = grp
            all_records.extend([r for grp_recs in recs.values() for r in grp_recs])

    if not all_records:
        return pd.DataFrame()

    df = pd.DataFrame(all_records)
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df = df.dropna(subset=["date"])
    return df.sort_values("date").reset_index(drop=True)


# ══════════════════════════════════════════════════════════════════════════════
# INVOICE PARSER
# ══════════════════════════════════════════════════════════════════════════════

def _get_tier(dish_name: str):
    for t in ["LOWCAL", "LEAN", "BUILD"]:
        if dish_name.upper().endswith(t):
            return t
    return None


def _bench_flag(meal_type: str, tier, price_ex: float) -> str:
    if not tier:
        return "NO_TIER"
    key = (meal_type, tier)
    min_b = BENCH_MIN.get(key)
    max_b = BENCH_MAX.get(key)
    if min_b is None:
        return "UNKNOWN"
    if price_ex > max_b:
        return "EXCEEDS ALL"
    if price_ex > min_b:
        return "CHECK"
    return "OK"


def parse_invoice_sheet(ws, year: int, month_label: str) -> list:
    """Parse one monthly invoice sheet. Returns list of dicts."""
    rows = list(ws.iter_rows(values_only=True))
    records = []

    for wk in INVOICE_WEEK_COLS:
        dc = wk["dish_col"]  - 1
        tc = wk["type_col"]  - 1
        qc = wk["qty_col"]   - 1
        pc = wk["price_col"] - 1
        current_date = None

        for row in rows[2:]:
            dish_val  = row[dc] if dc < len(row) else None
            row_type  = str(row[tc]).strip() if (tc < len(row) and row[tc]) else ""
            qty_val   = row[qc] if qc < len(row) else None
            price_val = row[pc] if pc < len(row) else None
            dish_name = str(dish_val).strip() if dish_val is not None else ""

            qty_str = str(qty_val).strip() if qty_val is not None else ""

            # Day total marker
            if qty_str in ("Total ", "Total"):
                if re.match(r"^[A-Z]{3}\s+\d{2}$", dish_name):
                    try:
                        current_date = datetime.strptime(
                            f"{dish_name} {year}", "%b %d %Y"
                        ).date()
                    except Exception:
                        pass
                continue

            # Skip compensation rows
            if any(kw in dish_name.lower() for kw in COMP_KW):
                continue
            if any(kw in row_type.lower() for kw in COMP_KW):
                continue

            # Skip headers, empty, out-of-scope types
            if dish_name in SKIP_DISHES or not dish_name or dish_name == "nan":
                continue
            if row_type in SKIP_TYPES:
                continue
            if row_type not in TYPE_MAP:
                continue
            if current_date is None:
                continue

            try:
                qty         = float(qty_val)   if qty_val   is not None else 0
                price_incl  = float(price_val) if price_val is not None else 0
            except (ValueError, TypeError):
                continue

            if qty <= 0 or price_incl <= 0:
                continue

            price_ex  = round(price_incl / VAT_DIVISOR, 4)
            meal_type = TYPE_MAP[row_type]
            tier      = _get_tier(dish_name)

            records.append({
                "date":                 current_date,
                "month":                month_label,
                "week":                 wk["week"],
                "meal_type":            meal_type,
                "dish_name":            dish_name,
                "tier":                 tier,
                "qty":                  qty,
                "unit_price_incl_vat":  round(price_incl, 4),
                "unit_price_ex_vat":    price_ex,
                "line_total_ex_vat":    round(qty * price_ex, 4),
                "bench_flag":           _bench_flag(meal_type, tier, price_ex),
            })

    return records


def load_all_invoices() -> pd.DataFrame:
    """
    Scan 02 Invoices folder for Excel files.
    Dynamically detects monthly sheet names (Daily Order-XXXXX).
    Returns flat DataFrame of all invoice lines.
    """
    invoice_dir = PATHS["invoices"]
    all_records = []

    for root, dirs, files in os.walk(invoice_dir):
        for fname in sorted(files):
            if not fname.endswith(".xlsx"):
                continue
            filepath = os.path.join(root, fname)
            try:
                wb = load_workbook(filepath, read_only=True)
            except Exception as e:
                print(f"  [WARN] Could not open {fname}: {e}")
                continue

            # Find all Daily Order sheets dynamically
            order_sheets = [s for s in wb.sheetnames
                            if s.startswith("Daily Order-") and s != "Daily Order-Format"]

            if not order_sheets:
                print(f"  [WARN] No Daily Order sheets in {fname}")
                print(f"         Sheets found: {wb.sheetnames}")
                continue

            for sheet_name in order_sheets:
                # Extract year from sheet (e.g. Daily Order-APR26 → 2026)
                year = 2026  # default
                ws = wb[sheet_name]
                month_label = sheet_name.replace("Daily Order-", "").replace("26", "")
                recs = parse_invoice_sheet(ws, year, month_label)
                if recs:
                    all_records.extend(recs)
                    print(f"  Parsed {sheet_name}: {len(recs)} lines")
                else:
                    print(f"  [WARN] No data parsed from {sheet_name}")

    if not all_records:
        return pd.DataFrame()

    df = pd.DataFrame(all_records)
    df["date"] = pd.to_datetime(df["date"])
    return df.sort_values("date").reset_index(drop=True)


# ══════════════════════════════════════════════════════════════════════════════
# ORDERS LOADER
# ══════════════════════════════════════════════════════════════════════════════

def load_orders() -> pd.DataFrame:
    """Load and enrich customer orders from reference folder."""
    ref_dir = PATHS["reference"]
    # Look for any xlsx file in reference folder
    order_files = glob.glob(os.path.join(ref_dir, "*.xlsx"))
    if not order_files:
        raise FileNotFoundError("No Excel file found in 03 Reference")

    df = pd.read_excel(order_files[0])
    df["full_name"] = (
        df["Name"].astype(str).str.strip() + " " +
        df["Surname"].astype(str).str.strip()
    ).str.strip()

    # Keep 100% discount orders — route to classpass or complimentary groups
    df = df.copy()

    df["Order date from"] = pd.to_datetime(df["Order date from"])
    df["Order date to"]   = pd.to_datetime(df["Order date to"])
    df["name_norm"]       = df["full_name"].apply(_normalize)
    df["is_hp"]   = df["Name"].astype(str).str.strip().str.upper().str.startswith("HP")
    df["is_val"]  = df["full_name"].str.contains("VALOREM", case=False, na=False)
    df["is_cp"]   = df["full_name"].str.strip().str.upper().str.endswith(" CP")
    df["is_comp"] = (
        (df["Percentage discount"].astype(str).str.strip() == "100.0%") &
        (~df["is_hp"]) & (~df["is_val"]) & (~df["is_cp"])
    )
    df["Deposits"]        = pd.to_numeric(df["Deposits"], errors="coerce").fillna(0)
    df["paid_ex_vat"]     = (df["Gross value"].astype(float) / VAT_DIVISOR).round(2)
    df["net_revenue"]     = (df["paid_ex_vat"] - df["Deposits"]).round(2)

    return df
