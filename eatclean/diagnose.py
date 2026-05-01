"""
EatClean Pipeline Diagnostics
Usage: python3 diagnose.py

Shows exactly what the pipeline does for a specific client on a specific date:
- What dishes were parsed from the Nutribot file
- What price was found for each dish
- Why a dish may be missing or wrongly priced

Run from the eatclean folder.
"""

import os, re, sys, glob, tempfile
import pandas as pd
from openpyxl import load_workbook

# ── CONFIG ────────────────────────────────────────────────────────────────────
SEARCH_NAME = input("Enter client name (partial match OK): ").strip().lower()
SEARCH_DATE = input("Enter date (YYYY-MM-DD, e.g. 2026-04-09): ").strip()

# ── DRIVE PATH ────────────────────────────────────────────────────────────────
DRIVE_ROOT = os.path.join(
    os.path.expanduser("~"), "Library", "CloudStorage",
    "GoogleDrive-admin@eatcleanme.com", "My Drive", "02 Automation"
)
SEL_DIR = os.path.join(DRIVE_ROOT, "01 Daily_selections")
INV_DIR = os.path.join(DRIVE_ROOT, "02 Invoices")

print(f"\n{'='*60}")
print(f"DIAGNOSTIC: '{SEARCH_NAME}' on {SEARCH_DATE}")
print(f"{'='*60}")

# ── STEP 1: BUILD PRICE LOOKUP ────────────────────────────────────────────────
print("\n[1] Building price lookup from invoice Price List...")

from config import MANUAL_PRICES, VAT_DIVISOR

lookup = {}
inv_files = glob.glob(os.path.join(INV_DIR, "*.xlsx"))
if not inv_files:
    inv_files = glob.glob(os.path.join(INV_DIR, "**", "*.xlsx"), recursive=True)

if inv_files:
    price_file = inv_files[0]
    print(f"    Using: {os.path.basename(price_file)}")
    try:
        wb = load_workbook(price_file, read_only=True, data_only=True)
        if "Price List" in wb.sheetnames:
            ws = wb["Price List"]
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0: continue
                if not row[0] or not row[1]: continue
                try:
                    name  = str(row[0]).strip()
                    price = float(row[1]) / VAT_DIVISOR
                    if name and price > 0:
                        lookup[name] = round(price, 4)
                except: continue
            print(f"    ✅ {len(lookup)} dishes loaded from Price List sheet")
        else:
            print(f"    ⚠️  No 'Price List' sheet found. Sheets: {wb.sheetnames}")
    except Exception as e:
        print(f"    ❌ Error: {e}")
else:
    print("    ❌ No invoice file found in 02 Invoices")

# Apply manual overrides
for k, v in MANUAL_PRICES.items():
    if k not in lookup:
        lookup[k] = v
print(f"    Total with manual overrides: {len(lookup)} dishes")

# ── STEP 2: FIND THE DAILY FILE ────────────────────────────────────────────────
print(f"\n[2] Finding daily selection file for {SEARCH_DATE}...")

# Parse date parts
try:
    y, m, d = SEARCH_DATE.split("-")
    month_folder = f"{y}-{m}"
except:
    print("    ❌ Invalid date format. Use YYYY-MM-DD")
    sys.exit(1)

# Look for standard format file
standard_pattern = os.path.join(SEL_DIR, month_folder,
                                f"ordered-meal-plans-{SEARCH_DATE}.xlsx")
report_pattern   = os.path.join(SEL_DIR, month_folder,
                                f"Ordered_Meal_Plans_Report_{d}-{m}-{y}.xlsx")

found_file = None
file_type  = None

if os.path.exists(standard_pattern):
    found_file = standard_pattern
    file_type  = "standard"
elif os.path.exists(report_pattern):
    found_file = report_pattern
    file_type  = "report"
else:
    # Search all files in month folder
    month_dir = os.path.join(SEL_DIR, month_folder)
    if os.path.exists(month_dir):
        all_files = sorted(os.listdir(month_dir))
        print(f"    Files in {month_folder}: {all_files}")
    print(f"    ❌ No file found for {SEARCH_DATE}")
    print(f"    Looked for: {os.path.basename(standard_pattern)}")
    print(f"    Looked for: {os.path.basename(report_pattern)}")
    sys.exit(1)

print(f"    ✅ Found [{file_type}]: {os.path.basename(found_file)}")

# ── STEP 3: PARSE FILE & FIND CLIENT ──────────────────────────────────────────
print(f"\n[3] Parsing file and finding '{SEARCH_NAME}'...")

COLS_10 = ["customer_id","person","address","meal_plan","calories",
           "meals","exclusions","order_date","order_period","nominal_price"]
COLS_9  = ["customer_id","person","address","meal_plan","calories",
           "meals","exclusions","order_date","nominal_price"]

client_rows = []

if file_type == "standard":
    df = pd.read_excel(found_file, header=0)
    print(f"    Total rows in file: {len(df)}")
    print(f"    Columns: {list(df.columns)}")
    for _, row in df.iterrows():
        person = str(row.get("Person", "")).strip()
        if SEARCH_NAME in person.lower():
            client_rows.append({
                "person": person,
                "date":   str(row.get("Meal plan for the day", ""))[:10],
                "meals":  str(row.get("Meals", "")),
            })

elif file_type == "report":
    sheets = pd.ExcelFile(found_file).sheet_names
    for s in sheets:
        df_s = pd.read_excel(found_file, sheet_name=s, header=0)
        if len(df_s.columns) == 10:
            df_s.columns = COLS_10
        else:
            df_s.columns = COLS_9
        for _, row in df_s.iterrows():
            person = str(row.get("person", "")).strip()
            if SEARCH_NAME in person.lower():
                client_rows.append({
                    "person": person,
                    "date":   SEARCH_DATE,
                    "meals":  str(row.get("meals", "")),
                })

if not client_rows:
    print(f"    ❌ '{SEARCH_NAME}' not found in this file")
    # Show all persons in file
    if file_type == "standard":
        df = pd.read_excel(found_file, header=0)
        persons = sorted(df["Person"].dropna().unique())
    else:
        df_s = pd.read_excel(found_file, sheet_name=0, header=0)
        df_s.columns = COLS_10 if len(df_s.columns) == 10 else COLS_9
        persons = sorted(df_s["person"].dropna().unique())
    print(f"\n    All persons in this file ({len(persons)}):")
    for p in persons:
        print(f"      '{p}'")
    sys.exit(0)

print(f"    ✅ Found {len(client_rows)} row(s) for '{SEARCH_NAME}'")

# ── STEP 4: PARSE DISHES AND PRICES ───────────────────────────────────────────
print(f"\n[4] Dish-by-dish breakdown:")
print(f"{'─'*70}")

total_cost = 0
for cr in client_rows:
    print(f"\n  Person: {cr['person']}  |  Date: {cr['date']}")
    print(f"  {'Meal Type':<14} {'Dish Name':<40} {'Price':>10} {'Status'}")
    print(f"  {'─'*14} {'─'*40} {'─'*10} {'─'*20}")

    raw_meals = cr["meals"]
    day_cost  = 0
    found     = 0
    missing   = 0

    for line in raw_meals.split("\n"):
        line = line.strip()
        if not line or ":" not in line:
            continue
        meal_type, dish_raw = line.split(":", 1)
        dish_raw  = dish_raw.strip().replace("(x)", "").strip()
        mult_match = re.match(r"^(\d+)x\s+", dish_raw)
        mult = int(mult_match.group(1)) if mult_match else 1
        if mult_match:
            dish_raw = dish_raw[mult_match.end():]
        dish_name = re.sub(r"\s*-\s*\d+\s*$", "", dish_raw).strip()
        dish_name = re.sub(r"\s+", " ", dish_name)
        if not dish_name or dish_name == "nan":
            continue

        is_gram = bool(re.match(r"^\d+(?:\.\d+)?g\s", dish_name))
        price   = lookup.get(dish_name, 0)
        line_total = round(price * mult, 3)
        day_cost  += line_total

        if is_gram:
            status = "⏭ GRAM-BASED (skipped)"
        elif price > 0:
            status = f"✅ AED {price:.3f}"
            found += 1
        else:
            status = "❌ NOT IN PRICE LIST"
            missing += 1

        qty_str = f"×{mult}" if mult > 1 else ""
        print(f"  {meal_type.strip():<14} {dish_name:<40} {qty_str:>4} {f'AED {line_total:.3f}':>10}  {status}")

    print(f"\n  {'─'*70}")
    print(f"  Day total kitchen cost: AED {day_cost:.3f}")
    print(f"  Priced: {found} dishes | ❌ Missing: {missing} dishes")
    total_cost += day_cost

print(f"\n{'='*60}")
print(f"SUMMARY FOR '{SEARCH_NAME}' ON {SEARCH_DATE}")
print(f"  Total kitchen cost: AED {total_cost:.2f}")
print(f"{'='*60}")

# ── STEP 5: SHOW MISSING DISHES ────────────────────────────────────────────────
missing_dishes = []
for cr in client_rows:
    for line in cr["meals"].split("\n"):
        line = line.strip()
        if not line or ":" not in line: continue
        _, dish_raw = line.split(":", 1)
        dish_raw  = dish_raw.strip().replace("(x)", "").strip()
        mult_match = re.match(r"^(\d+)x\s+", dish_raw)
        if mult_match: dish_raw = dish_raw[mult_match.end():]
        dish_name = re.sub(r"\s*-\s*\d+\s*$", "", dish_raw).strip()
        dish_name = re.sub(r"\s+", " ", dish_name)
        if not dish_name or re.match(r"^\d+g\s", dish_name): continue
        if lookup.get(dish_name, 0) == 0:
            missing_dishes.append(dish_name)

if missing_dishes:
    print(f"\n[5] Missing dishes (not in price list):")
    for d in missing_dishes:
        # Try to find close matches
        close = [k for k in lookup if d.lower()[:10] in k.lower()][:3]
        print(f"  ❌ '{d}'")
        if close:
            print(f"     Closest in price list: {close}")
